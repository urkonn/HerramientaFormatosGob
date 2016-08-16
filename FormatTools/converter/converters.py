"""
Clases y metodos para manejo y conversion
de archivos xls y xlsx a formatos abiertos:
csv, json, txt, xml, html
"""
import os
import csv
import json
import xlrd
import datetime
from openpyxl import load_workbook
from django.conf import settings


def save_temporary_xls(xls_file):
    """
    Metodo que guarda el xls o xlsx
    de forma temporal en el sistema de archivos
    """
    base_path = os.path.join(settings.TEMPORAL_FILES_ROOT, "xls/")
    xls_file_path = os.path.join(base_path, limpia_nombre_archivo(xls_file.name))

    with open(xls_file_path.encode('utf-8'), 'w') as temporal_file:
        temporal_file.write(xls_file.read())

    return xls_file_path


def limpia_nombre_archivo(archivo):
    """
    Funcion que devuelve el nombre
    de un archivo sin formato
    ni extensiones
    Return: String
    """
    name = archivo.split('/')
    return name[(len(name) - 1)]


def encode_row(row, txt_convert=False, is_xlsx=False):
    """
    Funcion que valida el encoding de un valor
    para manejo del utf-8 en la clase csv
    Return String o Unicode
    """
    def encode_xls():
        """
        Funcion generadora que recorre
        cada celda de una fila y la
        transforma a Unicode con utf-8
        Solo para xls
        """
        for item in row:
            if txt_convert:
                yield item.encode('utf-8') if type(item) == type(u'') else str(item)
            else:
                yield item.encode('utf-8') if type(item) == type(u'') else item

    def encode_xlsx():
        """
        Funcion generadora que recorre
        cada celda de una fila y la
        transforma a Unicode con utf-8
        Solo para xlsx
        """
        for item in row:
            value = item.value
            if type(value) == datetime.datetime:
                yield value.strftime('%Y-%m-%dT%H:%M:%SZ')
            else:
                if txt_convert:
                    yield value.encode('utf-8') if type(value) == type(u'') else str(value)
                else:
                    yield value.encode('utf-8') if type(value) == type(u'') else value

    return encode_xlsx() if is_xlsx else encode_xls()


class XLSConverter(object):
    """
    Clase interfaz para acceder
    a los convertidores y sus
    propiedades estaticas
    """
    @staticmethod
    def get_converter(format_file, memory_file, xls_name=None):
        """ Funcion que retorna la instancia del convertidor
        dependiendo del formato y asociada al archivo en memoria
        Tipo de Retorno: XLSConverter or None
        """
        for decoder in XLSConverterBase.__subclasses__():
            if decoder.get_format_type() == format_file:
                return decoder(memory_file=memory_file, xls_name=xls_name)
        return None

    @staticmethod
    def get_mime_type_of_file(format_file):
        """
        Funcion que retorna el MIME type
        dependiendo del formato de archivo
        Tipo de Retorno: String
        """
        for decoder in XLSConverterBase.__subclasses__():
            if decoder.get_format_type() == format_file:
                return decoder.get_mime_type()


class XLSConverterBase(object):
    """ Clase base de los convertidores
    contiene metodos comunes de todos los convertidores
    """
    def __init__(self, *args, **kwargs):
        """ Constructor de la clase:
        Agrega como parametro el archivo en memoria (xls, xlsx)
        Obtiene la hoja de trabajo principal
        """
        self.memory_file = kwargs.get('memory_file', None)
        self.xls_name = kwargs.get('xls_name', None)

        self.load_temporal_file()
        self.is_xlsx = True if '.xlsx' in self.memory_file.name else False

        if self.is_xlsx:
            self.xls_file = load_workbook(self.get_path_temporal_xls(), use_iterators=True, read_only=True)
            self.principal_sheet = self.xls_file.get_sheet_by_name(self.xls_file.sheetnames[0])
        else:
            self.xls_file = xlrd.open_workbook(self.get_path_temporal_xls(), encoding_override="cp1252", on_demand=True)
            self.principal_sheet = self.xls_file.sheet_by_index(0)

    def load_temporal_file(self):
        """
        Metodo que carga en memoria
        el archivo xls o xlsx temporal
        """
        with open(self.memory_file) as temporal_xls:
            self.memory_file = temporal_xls

    def get_rows_of_xls(self):
        """ Funcion que devuelve una funcion generadora
        para recorrer todas las filas de la hoja de trabajo
        Tipo de Retorno: Function
        """
        if not self.xls_file:
            return []

        def xls_row_generator():
            """
            Funcion generadora que
            devuelve todas las filas
            de un xls o xlsx
            """
            if not self.is_xlsx:
                for index_row in xrange(self.principal_sheet.nrows):
                    yield self.principal_sheet.row_values(index_row)

        return self.principal_sheet.rows if self.is_xlsx else xls_row_generator()

    def get_total_of_rows(self):
        """ Funcion que devuelve el numero de total
        de filas dentro del archivo xls
        Tipo de Retorno: Function
        """
        return self.principal_sheet.max_row if self.is_xlsx else self.principal_sheet.nrows

    def get_name_file(self):
        """ Funcion que devuelve una funcion generadora
        para recorrer todas las filas de la hoja de trabajo
        Tipo de Retorno: Function
        """
        return '{0}.{1}'.format(self.get_clean_name_file(), self.get_format_type())

    def get_clean_name_file(self):
        """ Funcion que devuelve una funcion generadora
        para recorrer todas las filas de la hoja de trabajo
        Tipo de Retorno: Function
        """
        name = self.xls_name.replace('.xlsx', '').replace('.xls', '')
        return name

    def get_path_temporal_xls(self):
        """ Funcion que devuelve el path donde se encuentra temporalmente alojado
        el archivo xls
        Tipo de Retorno: String
        """
        base_path = os.path.join(settings.TEMPORAL_FILES_ROOT, "xls/")
        return os.path.join(base_path, self.memory_file.name)

    def get_path_file(self):
        """ Funcion que devuelve el path donde se encuentra temporalmente alojado
        el archivo resultante del proceso de conversion
        Tipo de Retorno: String
        """
        base_path = os.path.join(settings.TEMPORAL_FILES_ROOT, self.get_format_type())
        return os.path.join(base_path, self.get_name_file())

    def get_download_link(self):
        """
        Funcion que devuelve el link de descarga del archivo
        """
        return '{0}converter/download/{1}/{2}'.format(settings.FQDN, self.get_format_type(), self.get_name_file())

    def convert(self):
        """ Funcion que convierte el archivo xls
        al tipo especifico del convertidor, genera un archivo temporal
        y retorna un link de descarga
        Tipo de Retorno: URL
        """
        pass

    @staticmethod
    def get_mime_type():
        """ Funcion que devuelve el MIME type del convertidor
        dependiendo del tipo de archivo que maneja
        Tipo de Retorno: String
        """
        pass

    @staticmethod
    def get_format_type():
        """
        Funcion estatica que devuelve
        el tipo de archivo manejado
        por el convertidor
        """
        pass


class XLSToCSVConverter(XLSConverterBase):
    """ Clase Convertidor especializada
    en la transformacion de formatos xls y xlsx a csv
    """
    def __init__(self, *args, **kwargs):
        self.format_extension = 'csv'
        self.csv_writer = None
        super(XLSToCSVConverter, self).__init__(*args, **kwargs)

    @staticmethod
    def get_format_type():
        """
        Funcion estatica que devuelve
        el tipo de archivo manejado
        por el convertidor
        """
        return 'csv'

    def convert(self):
        """ Funcion que convierte el archivo xls
        al tipo especifico del convertidor, genera un archivo temporal
        y retorna un link de descarga
        Tipo de Retorno: URL
        """
        def converter_generator():
            """
            Funcion generadora que
            lleva a cabo la tarea de
            conversion
            """
            rows_count = self.get_total_of_rows()
            counter = 1
            with open(self.get_path_file(), 'w') as csv_file:
                self.csv_writer = csv.writer(csv_file)
                for row in self.get_rows_of_xls():
                    self.csv_writer.writerow(list(encode_row(row, is_xlsx=self.is_xlsx)))
                    yield '%.2f' % ((counter /float(rows_count)) * 100)
                    counter += 1

        return converter_generator()

    @staticmethod
    def get_mime_type():
        """ Funcion que devuelve el MIME type del convertidor
        dependiendo del tipo de archivo que maneja
        Tipo de Retorno: String
        """
        return 'text/csv'


class XLSToJSONConverter(XLSConverterBase):
    """
    Clase Convertidor especializada
    en la transformacion de formatos
    xls y xlsx a json
    """
    def __init__(self, *args, **kwargs):
        self.format_extension = 'json'
        super(XLSToJSONConverter, self).__init__(*args, **kwargs)

    @staticmethod
    def get_format_type():
        """
        Funcion estatica que devuelve
        el tipo de archivo manejado
        por el convertidor
        """
        return 'json'

    def convert(self):
        """ Funcion que convierte el archivo xls
        al tipo especifico del convertidor, genera un archivo temporal
        y retorna un link de descarga
        Tipo de Retorno: URL
        """
        def converter_generator():
            """
            Funcion generadora que
            lleva a cabo la tarea de
            conversion
            """
            index_row = 0
            separator = ''
            rows_count = self.get_total_of_rows()

            with open(self.get_path_file(), 'w') as json_file:
                json_file.write("{{name: \"{0}\", count_rows: {1}, rows: [".format(self.get_clean_name_file(), self.principal_sheet.max_row if self.is_xlsx else self.principal_sheet.nrows))

                for row in self.get_rows_of_xls():
                    json_file.write("{0}{1}".format(separator, json.dumps({'row': index_row, 'values': list(encode_row(row, is_xlsx=self.is_xlsx))})))

                    if index_row == 0:
                        separator = ','

                    index_row += 1
                    yield '%.2f' % ((index_row /float(rows_count)) * 100)

                max_rows = self.principal_sheet.max_row if self.is_xlsx else self.principal_sheet.nrows
                if max_rows == index_row:
                    json_file.write("]}")

        return converter_generator()

    @staticmethod
    def get_mime_type():
        """ Funcion que devuelve el MIME type del convertidor
        dependiendo del tipo de archivo que maneja
        Tipo de Retorno: String
        """
        return 'application/json'


class XLSToTxtConverter(XLSConverterBase):
    """ Clase Convertidor especializada
    en la transformacion de formatos xls y xlsx a txt
    """
    def __init__(self, *args, **kwargs):
        self.format_extension = 'txt'
        super(XLSToTxtConverter, self).__init__(*args, **kwargs)

    @staticmethod
    def get_format_type():
        """
        Funcion estatica que devuelve
        el tipo de archivo manejado
        por el convertidor
        """
        return 'txt'

    def convert(self):
        """ Funcion que convierte el archivo xls
        al tipo especifico del convertidor, genera un archivo temporal
        y retorna un link de descarga
        Tipo de Retorno: URL
        """
        def convert_generator():
            """
            Funcion generadora que
            lleva a cabo la tarea de
            conversion
            """
            rows_count = self.get_total_of_rows()
            counter = 1
            with open(self.get_path_file(), 'w') as txt_file:
                for row in self.get_rows_of_xls():
                    txt_file.write('{0}\n'.format(';'.join(list(encode_row(row, txt_convert=True, is_xlsx=self.is_xlsx)))))
                    yield '%.2f' % ((counter /float(rows_count)) * 100)
                    counter += 1

        return convert_generator()

    @staticmethod
    def get_mime_type():
        """ Funcion que devuelve el MIME type del convertidor
        dependiendo del tipo de archivo que maneja
        Tipo de Retorno: String
        """
        return 'text/plain'


class XLSToXMLConverter(XLSConverterBase):
    """ Clase Convertidor especializada
    en la transformacion de formatos xls y xlsx a xml
    """
    def __init__(self, *args, **kwargs):
        self.format_extension = 'xml'
        super(XLSToXMLConverter, self).__init__(*args, **kwargs)

    @staticmethod
    def get_format_type():
        """
        Funcion estatica que devuelve
        el tipo de archivo manejado
        por el convertidor
        """
        return 'xml'

    def convert(self):
        """ Funcion que convierte el archivo xls
        al tipo especifico del convertidor, genera un archivo temporal
        y retorna un link de descarga
        Tipo de Retorno: URL
        """
        def convert_generator():
            """
            Funcion generadora que
            lleva a cabo la tarea de
            conversion
            """
            rows_count = self.get_total_of_rows()
            counter = 1
            with open(self.get_path_file(), 'w') as xml_file:
                xml_file.write("<?xml version=\"1.0\" encoding=\"UTF-8\"?><rows>")
                for row in self.get_rows_of_xls():
                    xml_file.write("<row>")
                    for field in encode_row(row, txt_convert=True, is_xlsx=self.is_xlsx):
                        xml_file.write("<field>{0}</field>".format(field))
                    xml_file.write("</row>")

                yield '%.2f' % ((counter /float(rows_count)) * 100)
                counter += 1
                xml_file.write("</rows>")

        return convert_generator()

    @staticmethod
    def get_mime_type():
        """ Funcion que devuelve el MIME type del convertidor
        dependiendo del tipo de archivo que maneja
        Tipo de Retorno: String
        """
        return 'application/xml'


class XLSToHTMLConverter(XLSConverterBase):
    """ Clase Convertidor especializada
    en la transformacion de formatos xls y xlsx a HTML
    """
    def __init__(self, *args, **kwargs):
        self.format_extension = 'html'
        super(XLSToHTMLConverter, self).__init__(*args, **kwargs)

    @staticmethod
    def get_format_type():
        """
        Funcion estatica que devuelve
        el tipo de archivo manejado
        por el convertidor
        """
        return 'html'

    def convert(self):
        """ Funcion que convierte el archivo xls
        al tipo especifico del convertidor, genera un archivo temporal
        y retorna un link de descarga
        Tipo de Retorno: URL
        """
        def convert_generator():
            """
            Funcion generadora que
            lleva a cabo la tarea de
            conversion
            """
            rows_count = self.get_total_of_rows()
            counter = 1
            with open(self.get_path_file(), 'w') as html_file:
                html_file.write("<meta charset='utf-8'><table attr-name='{0}'>".format(self.get_clean_name_file()))

                for row in self.get_rows_of_xls():
                    html_file.write("<tr>")
                    for field in encode_row(row, txt_convert=True, is_xlsx=self.is_xlsx):
                        html_file.write("<td>{0}</td>".format(field))
                    html_file.write("</tr>")

                    yield '%.2f' % ((counter /float(rows_count)) * 100)
                    counter += 1
                html_file.write("</table>")

        return convert_generator()

    @staticmethod
    def get_mime_type():
        """ Funcion que devuelve el MIME type del convertidor
        dependiendo del tipo de archivo que maneja
        Tipo de Retorno: String
        """
        return 'text/html'
